import openpyxl
from pathlib import Path
import csv

DATA_DIR = Path("data")

FILES = [
    "Cisco Unit List Price_240426.xlsx",
    "Devices Specs_240426.xlsx",
    "Network Quotation Tool_210426.xlsx",
]

OUTPUT_FILE = "formula_map.csv"


def inspect_workbook(file_path: Path):
    wb = openpyxl.load_workbook(file_path, data_only=False)

    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value

                if value is None:
                    continue

                is_formula = isinstance(value, str) and value.startswith("=")

                rows.append({
                    "file": file_path.name,
                    "sheet": sheet_name,
                    "cell": cell.coordinate,
                    "value": value,
                    "is_formula": is_formula,
                })

    return rows


def main():
    all_rows = []

    for file_name in FILES:
        file_path = DATA_DIR / file_name

        if not file_path.exists():
            print(f"missing file: {file_path}")
            continue

        print(f"reading: {file_path}")
        all_rows.extend(inspect_workbook(file_path))

    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["file", "sheet", "cell", "value", "is_formula"]
        )
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"done. output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()