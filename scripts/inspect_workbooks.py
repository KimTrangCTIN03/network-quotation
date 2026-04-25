import openpyxl
from pathlib import Path

DATA_DIR = Path("data")

FILES = [
    "Cisco Unit List Price_240426.xlsx",
    "Devices Specs_240426.xlsx",
    "Network Quotation Tool_210426 (1).xlsx",
]


def main():
    for file_name in FILES:
        file_path = DATA_DIR / file_name

        print("=" * 80)
        print(file_name)

        wb = openpyxl.load_workbook(file_path, data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            print("-" * 80)
            print("SHEET:", sheet_name)
            print("MAX ROW:", ws.max_row)
            print("MAX COL:", ws.max_column)

            print("FIRST 10 ROWS:")

            for r in range(1, min(ws.max_row, 10) + 1):
                values = []
                for c in range(1, min(ws.max_column, 12) + 1):
                    values.append(ws.cell(r, c).value)
                print(r, values)


if __name__ == "__main__":
    main()