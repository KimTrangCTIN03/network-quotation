import sys
import re
import pprint
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import range_boundaries


ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR))

from app.catalog_engine import cell_value, normalize_proposal_key


DATA_DIR = ROOT_DIR / "data"
TOOL_FILE = DATA_DIR / "Network Quotation Tool_210426 (1).xlsx"
OUTPUT_FILE = ROOT_DIR / "app" / "proposal_option_map.py"


def clean_option_model(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if not text:
        return ""

    return text


def is_valid_option_model(value: Any) -> bool:
    if value is None:
        return False

    if isinstance(value, (int, float)):
        return False

    text = str(value).strip()

    if not text:
        return False

    upper = text.upper()
    lower = text.lower()

    invalid_values = {
        "0",
        "0.0",
        "N/A",
        "NA",
        "-",
        "NONE",
        "LOW END",
        "MID RANGE",
        "HIGH END",
        "OPT 1",
        "OPT 2",
        "OPT 3",
        "OPTION 1",
        "OPTION 2",
        "OPTION 3",
        "SỐ LƯỢNG",
        "SO LUONG",
    }

    if upper in invalid_values:
        return False

    if text.startswith("$"):
        return False

    if lower.startswith("giá"):
        return False

    if lower.startswith("please"):
        return False

    if re.fullmatch(r"\d+(\.\d+)?", text):
        return False

    return True


def option_map_empty() -> Dict[str, List[str]]:
    return {
        "Low End": [],
        "Mid Range": [],
        "High End": [],
    }


def cell_display_value(wb, ws, row: int, col: int) -> Any:
    value = cell_value(wb, ws, row, col)

    if value is not None:
        return value

    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return cell_value(wb, ws, merged_range.min_row, merged_range.min_col)

    return value


def normalize_option_class(value: Any) -> str:
    text = str(value or "").strip().lower()

    if not text:
        return ""

    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)

    if "low end" in text:
        return "Low End"

    if "mid range" in text or "mid-range" in text:
        return "Mid Range"

    if "high end" in text:
        return "High End"

    if re.fullmatch(r"(opt|option)\s*1", text):
        return "Low End"

    if re.fullmatch(r"(opt|option)\s*2", text):
        return "Mid Range"

    if re.fullmatch(r"(opt|option)\s*3", text):
        return "High End"

    return ""


def find_class_above_cell(wb, ws, row: int, col: int) -> str:
    nearby_cols = [
        col,
        col - 1,
        col + 1,
        col - 2,
        col + 2,
        col - 3,
        col + 3,
    ]

    for r in range(row - 1, max(1, row - 15), -1):
        for c in nearby_cols:
            if c < 1 or c > ws.max_column:
                continue

            value = cell_display_value(wb, ws, r, c)
            class_name = normalize_option_class(value)

            if class_name:
                return class_name

    return ""


def is_group_label(text: str) -> bool:
    normalized = str(text or "").strip().lower()

    if not normalized:
        return False

    return (
        "campus" in normalized
        or normalized == "server farm"
        or normalized == "wan"
    )


def find_item_label_left(wb, ws, row: int, col: int) -> tuple[str, int]:
    first_col_text = str(cell_display_value(wb, ws, row, 1) or "").strip()

    if first_col_text and not is_group_label(first_col_text):
        first_col_lower = first_col_text.lower()
        first_col_skip_values = {
            "sá»‘ lÆ°á»£ng",
            "so luong",
            "số lượng",
            "opt 1",
            "opt 2",
            "opt 3",
            "option 1",
            "option 2",
            "option 3",
            "low end",
            "mid range",
            "high end",
        }

        if (
            first_col_lower not in first_col_skip_values
            and not re.fullmatch(r"\d+(\.\d+)?", first_col_text)
            and not first_col_text.startswith("$")
        ):
            return first_col_text, 1

    for c in range(col - 1, max(1, col - 50), -1):
        value = cell_display_value(wb, ws, row, c)
        text = str(value or "").strip()

        if not text:
            continue

        lower = text.lower()

        skip_values = {
            "số lượng",
            "so luong",
            "opt 1",
            "opt 2",
            "opt 3",
            "option 1",
            "option 2",
            "option 3",
            "low end",
            "mid range",
            "high end",
        }

        if lower in skip_values:
            continue

        if is_group_label(text):
            continue

        if re.fullmatch(r"\d+(\.\d+)?", text):
            continue

        if text.startswith("$"):
            continue

        return text, c

    return "", 0


def find_group_for_item_row(wb, ws, row: int, label_col: int) -> str:
    for r in range(row, 0, -1):
        value = cell_display_value(wb, ws, r, label_col)
        text = str(value or "").strip()

        if not text:
            continue

        lower = text.lower()

        if "campus" in lower:
            return "Campus - Trụ sở chính"

        if lower == "server farm":
            return "Server Farm"

        if lower == "wan":
            return "WAN"

    return ""


def find_proposal_start_row(wb, ws) -> int:
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 50) + 1):
            value = cell_display_value(wb, ws, row, col)
            text = str(value or "").strip().upper()

            if "GIẢI PHÁP ĐỀ XUẤT" in text:
                return row

    return 1


def get_defined_name_values(wb, defined_name: str) -> List[str]:
    values: List[str] = []

    try:
        dn = wb.defined_names[defined_name]
    except Exception:
        try:
            dn = wb.defined_names.get(defined_name)
        except Exception:
            dn = None

    if not dn:
        return values

    try:
        destinations = list(dn.destinations)
    except Exception:
        return values

    for sheet_name, coord in destinations:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        try:
            min_col, min_row, max_col, max_row = range_boundaries(coord)
        except Exception:
            continue

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = cell_value(wb, ws, row, col)

                if not is_valid_option_model(value):
                    continue

                model = clean_option_model(value)

                if model not in values:
                    values.append(model)

    return values


def read_values_from_validation_formula(wb, base_ws, formula: Any) -> List[str]:
    if formula is None:
        return []

    text = str(formula).strip()

    if not text:
        return []

    if text.startswith("="):
        text = text[1:].strip()

    if text.startswith('"') and text.endswith('"'):
        result = []

        for item in text[1:-1].split(","):
            if is_valid_option_model(item):
                model = clean_option_model(item)

                if model not in result:
                    result.append(model)

        return result

    if "!" not in text and not re.search(r"\$?[A-Z]+\$?\d+", text):
        named_values = get_defined_name_values(wb, text)

        if named_values:
            return named_values

    sheet_name = base_ws.title
    coord = text

    if "!" in text:
        sheet_part, coord_part = text.rsplit("!", 1)
        sheet_name = sheet_part.strip().strip("'")
        coord = coord_part.strip()

    coord = coord.replace("$", "")

    if sheet_name not in wb.sheetnames:
        return []

    try:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
    except Exception:
        return []

    ws = wb[sheet_name]
    result = []

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            value = cell_value(wb, ws, row, col)

            if not is_valid_option_model(value):
                continue

            model = clean_option_model(value)

            if model not in result:
                result.append(model)

    return result


def iter_data_validation_cells(dv):
    ranges_obj = getattr(dv, "cells", None) or getattr(dv, "sqref", None)

    if not ranges_obj:
        return

    try:
        ranges = ranges_obj.ranges
    except Exception:
        ranges = str(ranges_obj).split()

    for cell_range in ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                yield row, col


def set_option_map_value(
    result: Dict[str, Dict[str, List[str]]],
    key: str,
    class_name: str,
    models: List[str],
):
    key = normalize_proposal_key(key)

    if not key or not class_name or not models:
        return

    if key not in result:
        result[key] = option_map_empty()

    result[key][class_name] = models


def build_proposal_option_map() -> Dict[str, Dict[str, List[str]]]:
    if not TOOL_FILE.exists():
        raise FileNotFoundError(f"Không thấy file: {TOOL_FILE}")

    wb = openpyxl.load_workbook(TOOL_FILE, data_only=False)

    if "Campus" not in wb.sheetnames:
        raise ValueError("Không thấy tab Campus trong file Network Quotation Tool")

    ws = wb["Campus"]

    if not ws.data_validations:
        raise ValueError("Tab Campus không có data validation dropdown")

    proposal_start_row = find_proposal_start_row(wb, ws)
    result: Dict[str, Dict[str, List[str]]] = {}

    for dv in ws.data_validations.dataValidation:
        if str(dv.type).lower() != "list":
            continue

        models = read_values_from_validation_formula(wb, ws, dv.formula1)

        if not models:
            continue

        for row, col in iter_data_validation_cells(dv):
            if row < proposal_start_row:
                continue

            class_name = find_class_above_cell(wb, ws, row, col)

            if not class_name:
                continue

            item_label, label_col = find_item_label_left(wb, ws, row, col)

            if not item_label:
                continue

            group_name = find_group_for_item_row(wb, ws, row, label_col)

            if not group_name:
                continue

            item_key = normalize_proposal_key(item_label)
            group_item_key = normalize_proposal_key(f"{group_name}||{item_label}")

            set_option_map_value(result, group_item_key, class_name, models)

            if item_key not in result:
                result[item_key] = option_map_empty()

            if not result[item_key].get(class_name):
                result[item_key][class_name] = models

    return result


def write_python_mapping(option_map: Dict[str, Dict[str, List[str]]]):
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    content = (
        "# File này được sinh tự động từ scripts/generate_proposal_option_map.py\n"
        "# Không sửa tay file này nếu chưa cần override logic Excel.\n\n"
        "PROPOSAL_OPTION_MAP = "
        + pprint.pformat(option_map, width=140, sort_dicts=False)
        + "\n"
    )

    OUTPUT_FILE.write_text(content, encoding="utf-8")


def main():
    option_map = build_proposal_option_map()
    write_python_mapping(option_map)

    print(f"Đã sinh file: {OUTPUT_FILE}")
    print(f"Số key mapping: {len(option_map)}")

    print("\n=== SERVER FARM ===")
    for key, value in option_map.items():
        if key.startswith("Server Farm||"):
            print(key, value)

    print("\n=== WAN ===")
    for key, value in option_map.items():
        if key.startswith("WAN||"):
            print(key, value)


if __name__ == "__main__":
    main()
