from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
TOOL_FILE = ROOT / "data" / "Network Quotation Tool_240426.xlsx"


def print_cells(sheet_name: str, start_row: int, end_row: int, start_col: int = 1, end_col: int = 8) -> None:
    wb = openpyxl.load_workbook(TOOL_FILE, data_only=False)
    ws = wb[sheet_name]

    print(f"Workbook: {TOOL_FILE.name}")
    print(f"Sheet: {sheet_name}")

    for row in range(start_row, end_row + 1):
        values = []

        for col in range(start_col, end_col + 1):
            value = ws.cell(row, col).value
            values.append("" if value is None else str(value))

        print(row, values)


def print_data_validations(sheet_name: str, start_row: int, end_row: int) -> None:
    wb = openpyxl.load_workbook(TOOL_FILE, data_only=False)
    ws = wb[sheet_name]

    print(f"Data validations: {sheet_name}!{start_row}:{end_row}")

    for validation in ws.data_validations.dataValidation:
        sqref = str(validation.sqref)

        if any(f"{col}{row}" in sqref for row in range(start_row, end_row + 1) for col in ["C", "E", "G"]):
            print(validation.type, validation.formula1, sqref)


if __name__ == "__main__":
    print_cells("Campus", 46, 57)
    print_data_validations("Campus", 47, 57)
