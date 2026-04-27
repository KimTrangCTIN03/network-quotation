from collections import OrderedDict
from io import BytesIO
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.catalog_engine import BOM_DIR, cell_value, is_integer_line_number, normalize_model, to_float
from app.catalog_engine import C8000_SECURE_LICENSE_PARTS, C8000_SECURE_VARIANTS
from app.pricing.rules import INDOOR_AP_POWER_INJECTOR, OUTDOOR_AP_ACCESSORIES, WIFI7_LICENSE_BUNDLE


OPTIONS = [
    ("opt1", "Option 1 - Low End"),
    ("opt2", "Option 2 - Mid Range"),
    ("opt3", "Option 3 - High End"),
]

DETAIL_BOM_SHEETS = [
    "C8000",
    "C9200",
    "C9300",
    "C1300",
    "C9500",
    "ACI",
    "SFP",
    "AP_input",
    "ISR1000_input",
    "C8000_secure_input",
]

DETAIL_BOM_FILES = [
    "C8000.xlsx",
    "C9200.xlsx",
    "C9300.xlsx",
    "C1300.xlsx",
    "C9500.xlsx",
    "ACI.xlsx",
    "SFP.xlsx",
    "AP.xlsx",
    "ISR1000.xlsx",
    "C8000_secure.xlsx",
]

BOM_COLUMN_DEFINITIONS = [
    ("line_number", "Line Number"),
    ("part_number", "Item Name"),
    ("smart_account_mandatory", "Smart Account Mandatory"),
    ("description", "Description"),
    ("group_name", "Group Name"),
    ("service_duration_months", "Service Duration (Months)"),
    ("estimated_lead_time_days", "Estimated Lead Time (Days)"),
    ("included_item", "Included Item"),
    ("total_quantity", "Quantity"),
    ("pricing_term", "Pricing Term"),
    ("list_price", "ListPrice"),
    ("extended_list_price", "Extended ListPrice"),
    ("discount_percent", "Discount %"),
    ("extended_selling_price", "Selling Price"),
    ("service_type", "Service Type"),
]
BOM_COLUMN_KEYS = [key for key, _label in BOM_COLUMN_DEFINITIONS]


def selected_base_model(model: str) -> str:
    return normalize_model(str(model or "").split("(", 1)[0].strip())


def selected_variant(model: str) -> str:
    text = str(model or "")

    if "(" not in text or ")" not in text:
        return ""

    return text.split("(", 1)[1].split(")", 1)[0].strip()


def infer_component_type(part_number: Any, line_number: Any = "") -> str:
    part = normalize_model(part_number).upper()
    line_text = str(line_number or "")

    if is_integer_line_number(line_number):
        return "Base"

    if part.startswith("CON-"):
        return "Support"

    if any(token in part for token in ["LIC", "DNA", "SDWAN", "SUB", "SEC", "STACK"]):
        return "License"

    if any(token in part for token in ["BRKT", "BRACKET", "MNT", "RM-"]):
        return "Bracket/Mount"

    if any(token in part for token in ["PWR", "CAB", "BLANK", "RFID", "INJ"]):
        return "Accessory"

    return "Component"


def normalized_candidates(model: str) -> List[str]:
    normalized = normalize_model(model)
    candidates = [normalized]

    for suffix in ["-A", "-ROW", "-RTG"]:
        candidates.append(normalize_model(f"{normalized}{suffix}"))

    return list(dict.fromkeys([c for c in candidates if c]))


def row_to_bom_part(
    wb,
    ws,
    row: int,
    selected_model: str,
    quote_quantity: float,
    option_key: str,
    line: Dict[str, Any],
    component_type: str = "",
    quantity_multiplier: float = 1,
) -> Dict[str, Any] | None:
    part_number = cell_value(wb, ws, row, 2)

    if not part_number:
        return None

    quantity_per_unit = to_float(cell_value(wb, ws, row, 9), 1) * quantity_multiplier
    list_price = to_float(cell_value(wb, ws, row, 11), 0)
    extended_list_price = to_float(cell_value(wb, ws, row, 12), list_price * to_float(cell_value(wb, ws, row, 9), 1)) * quantity_multiplier
    discount_percent = to_float(cell_value(wb, ws, row, 13), 0)
    selling_price = to_float(cell_value(wb, ws, row, 14), extended_list_price) * quantity_multiplier

    return {
        "option": option_key,
        "group": line.get("group", ""),
        "item_type": line.get("item_type", ""),
        "selected_model": selected_model,
        "component_type": component_type or infer_component_type(part_number, cell_value(wb, ws, row, 1)),
        "source_sheet": ws.title,
        "line_number": cell_value(wb, ws, row, 1),
        "part_number": str(part_number).strip(),
        "smart_account_mandatory": cell_value(wb, ws, row, 3),
        "description": cell_value(wb, ws, row, 4),
        "group_name": cell_value(wb, ws, row, 5),
        "service_duration_months": cell_value(wb, ws, row, 6),
        "estimated_lead_time_days": cell_value(wb, ws, row, 7),
        "included_item": cell_value(wb, ws, row, 8),
        "quantity_per_unit": quantity_per_unit,
        "quote_quantity": quote_quantity,
        "total_quantity": quantity_per_unit * quote_quantity,
        "pricing_term": cell_value(wb, ws, row, 10),
        "list_price": list_price,
        "extended_list_price": extended_list_price * quote_quantity,
        "discount_percent": discount_percent,
        "selling_price": selling_price,
        "extended_selling_price": selling_price * quote_quantity,
        "service_type": cell_value(wb, ws, row, 15),
    }


def find_part_block(
    workbooks: List[Any],
    model: str,
    selected_model: str,
    quote_quantity: float,
    option_key: str,
    line: Dict[str, Any],
    component_type: str = "",
    quantity_multiplier: float = 1,
) -> List[Dict[str, Any]]:
    candidates = set(normalized_candidates(model))

    for wb in workbooks:
        for sheet_name in DETAIL_BOM_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                line_number = cell_value(wb, ws, row, 1)
                part_number = cell_value(wb, ws, row, 2)

                if not is_integer_line_number(line_number) or normalize_model(part_number) not in candidates:
                    continue

                parts = []

                for part_row in range(row, ws.max_row + 1):
                    if part_row != row:
                        next_line_number = cell_value(wb, ws, part_row, 1)
                        next_part_number = cell_value(wb, ws, part_row, 2)

                        if is_integer_line_number(next_line_number) and next_part_number:
                            break

                    subtotal_label = str(cell_value(wb, ws, part_row, 13) or "").strip().lower()

                    if subtotal_label == "subtotal":
                        break

                    part = row_to_bom_part(
                        wb,
                        ws,
                        part_row,
                        selected_model,
                        quote_quantity,
                        option_key,
                        line,
                        component_type,
                        quantity_multiplier,
                    )

                    if part:
                        parts.append(part)

                return parts

    return []


def find_part_rows(
    workbooks: List[Any],
    model: str,
    selected_model: str,
    quote_quantity: float,
    option_key: str,
    line: Dict[str, Any],
    component_type: str,
    quantity_multiplier: float = 1,
) -> List[Dict[str, Any]]:
    candidates = set(normalized_candidates(model))
    parts = []

    for wb in workbooks:
        for sheet_name in DETAIL_BOM_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                part_number = cell_value(wb, ws, row, 2)

                if normalize_model(part_number) not in candidates:
                    continue

                part = row_to_bom_part(
                    wb,
                    ws,
                    row,
                    selected_model,
                    quote_quantity,
                    option_key,
                    line,
                    component_type,
                    quantity_multiplier,
                )

                if part:
                    parts.append(part)

        if parts:
            return parts

    return parts


def find_bundle_parts(
    workbooks: List[Any],
    selected_model: str,
    quote_quantity: float,
    option_key: str,
    line: Dict[str, Any],
) -> List[Dict[str, Any]]:
    base_model = selected_base_model(selected_model)
    candidates = set(normalized_candidates(base_model or selected_model))

    for wb in workbooks:
        for sheet_name in DETAIL_BOM_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                line_number = cell_value(wb, ws, row, 1)
                part_number = cell_value(wb, ws, row, 2)

                if not is_integer_line_number(line_number):
                    continue

                if normalize_model(part_number) not in candidates:
                    continue

                parts = []
                stop_row = ws.max_row + 1

                for part_row in range(row, ws.max_row + 1):
                    if part_row != row:
                        next_line_number = cell_value(wb, ws, part_row, 1)
                        next_part_number = cell_value(wb, ws, part_row, 2)

                        if is_integer_line_number(next_line_number) and next_part_number:
                            break

                    subtotal_label = str(cell_value(wb, ws, part_row, 13) or "").strip().lower()

                    if subtotal_label == "subtotal":
                        stop_row = part_row
                        break

                    part = row_to_bom_part(wb, ws, part_row, selected_model, quote_quantity, option_key, line)

                    if part:
                        parts.append(part)

                if ws.title == "ISR1000_input":
                    marker_row = None
                    model_key = normalize_model(selected_model).lower()

                    for scan_row in range(stop_row + 1, min(ws.max_row, stop_row + 4) + 1):
                        marker = str(cell_value(wb, ws, scan_row, 3) or "").lower()

                        if "license" in marker and model_key in normalize_model(marker).lower():
                            marker_row = scan_row
                            break

                    if marker_row:
                        for part_row in range(marker_row + 1, ws.max_row + 1):
                            subtotal_label = str(cell_value(wb, ws, part_row, 13) or "").strip().lower()

                            if subtotal_label == "subtotal":
                                break

                            part = row_to_bom_part(wb, ws, part_row, selected_model, quote_quantity, option_key, line)

                            if part:
                                parts.append(part)

                parts.extend(add_mapped_components(workbooks, selected_model, quote_quantity, option_key, line, ws.title))
                return parts

    return []


def add_mapped_components(
    workbooks: List[Any],
    selected_model: str,
    quote_quantity: float,
    option_key: str,
    line: Dict[str, Any],
    source_sheet: str = "",
) -> List[Dict[str, Any]]:
    model = selected_base_model(selected_model)
    extras: List[Dict[str, Any]] = []

    if source_sheet == "AP_input":
        if model.startswith("CW917"):
            extras.extend(find_part_block(workbooks, INDOOR_AP_POWER_INJECTOR, selected_model, quote_quantity, option_key, line, "Accessory"))
            extras.extend(find_part_block(workbooks, WIFI7_LICENSE_BUNDLE, selected_model, quote_quantity, option_key, line, "License", 1 / 5))
        elif model.startswith("C9124") or model == "CW9163E":
            for part_number in OUTDOOR_AP_ACCESSORIES:
                component_type = "Bracket/Mount" if "MNT" in part_number else "Accessory"
                extras.extend(find_part_block(workbooks, part_number, selected_model, quote_quantity, option_key, line, component_type))
        elif model.startswith("C9136") or model.startswith("CW916"):
            extras.extend(find_part_block(workbooks, INDOOR_AP_POWER_INJECTOR, selected_model, quote_quantity, option_key, line, "Accessory"))

    variant = selected_variant(selected_model)

    if source_sheet == "C8000_secure_input" and model in C8000_SECURE_VARIANTS and variant:
        size, _variants = C8000_SECURE_VARIANTS[model]

        for part_number in C8000_SECURE_LICENSE_PARTS.get(size, {}).get(variant, []):
            extras.extend(find_part_rows(workbooks, part_number, selected_model, quote_quantity, option_key, line, "License"))

    return extras


def load_detail_bom_workbooks() -> List[Any]:
    workbooks = []

    if not BOM_DIR.exists():
        return workbooks

    for file_name in DETAIL_BOM_FILES:
        file_path = BOM_DIR / file_name

        if file_path.exists():
            workbooks.append(openpyxl.load_workbook(file_path, data_only=False))

    return workbooks


def fallback_selected_device_row(
    selected: Dict[str, Any],
    quote_quantity: float,
    option_key: str,
    line: Dict[str, Any],
) -> Dict[str, Any]:
    price = to_float(selected.get("price"), 0)
    model = selected.get("model", "")

    return {
        "option": option_key,
        "group": line.get("group", ""),
        "item_type": line.get("item_type", ""),
        "selected_model": model,
        "component_type": "Fallback",
        "source_sheet": selected.get("sheet", ""),
        "line_number": "",
        "part_number": model,
        "smart_account_mandatory": "",
        "description": "No detailed BOM found in imported BOM files",
        "group_name": "",
        "service_duration_months": "",
        "estimated_lead_time_days": "",
        "included_item": "",
        "quantity_per_unit": 1,
        "quote_quantity": quote_quantity,
        "total_quantity": quote_quantity,
        "pricing_term": "",
        "list_price": price,
        "extended_list_price": price * quote_quantity,
        "discount_percent": 0,
        "selling_price": price,
        "extended_selling_price": price * quote_quantity,
        "service_type": "",
    }


def aggregate_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: OrderedDict[str, Dict[str, Any]] = OrderedDict()

    for row in rows:
        key = (
            row.get("part_number", ""),
            row.get("description", ""),
            row.get("component_type", ""),
            row.get("selling_price", 0),
            row.get("list_price", 0),
        )
        key_text = repr(key)

        if key_text not in grouped:
            grouped[key_text] = dict(row)
            grouped[key_text]["selected_model"] = row.get("selected_model", "")
            grouped[key_text]["component_type"] = row.get("component_type", "")
            grouped[key_text]["item_type"] = row.get("item_type", "")
            grouped[key_text]["group"] = row.get("group", "")
            continue

        target = grouped[key_text]
        target["total_quantity"] = to_float(target.get("total_quantity"), 0) + to_float(row.get("total_quantity"), 0)
        target["extended_list_price"] = to_float(target.get("extended_list_price"), 0) + to_float(row.get("extended_list_price"), 0)
        target["extended_selling_price"] = to_float(target.get("extended_selling_price"), 0) + to_float(row.get("extended_selling_price"), 0)

        for field in ["selected_model", "component_type", "item_type", "group", "source_sheet"]:
            current_values = [v.strip() for v in str(target.get(field) or "").split(";") if v.strip()]
            new_value = str(row.get(field) or "").strip()

            if new_value and new_value not in current_values:
                current_values.append(new_value)

            target[field] = "; ".join(current_values)

    return list(grouped.values())


def group_header_row(option_key: str, group_name: str) -> Dict[str, Any]:
    return {
        "option": option_key,
        "group": group_name,
        "item_type": "",
        "selected_model": "",
        "component_type": "Group Header",
        "source_sheet": "",
        "line_number": "",
        "part_number": group_name,
        "smart_account_mandatory": "",
        "description": group_name,
        "group_name": "",
        "service_duration_months": "",
        "estimated_lead_time_days": "",
        "included_item": "",
        "quantity_per_unit": "",
        "quote_quantity": "",
        "total_quantity": "",
        "pricing_term": "",
        "list_price": "",
        "extended_list_price": "",
        "discount_percent": "",
        "selling_price": "",
        "extended_selling_price": "",
        "service_type": "",
        "is_group_header": True,
    }


def subtotal_row(option_key: str, group_name: str, block_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "option": option_key,
        "group": group_name,
        "item_type": "",
        "selected_model": "",
        "component_type": "SubTotal",
        "source_sheet": "",
        "line_number": "",
        "part_number": "",
        "smart_account_mandatory": "",
        "description": "",
        "group_name": "",
        "service_duration_months": "",
        "estimated_lead_time_days": "",
        "included_item": "",
        "quantity_per_unit": "",
        "quote_quantity": "",
        "total_quantity": "",
        "pricing_term": "",
        "list_price": "",
        "extended_list_price": sum(to_float(row.get("extended_list_price"), 0) for row in block_rows),
        "discount_percent": "SubTotal",
        "selling_price": "",
        "extended_selling_price": sum(to_float(row.get("extended_selling_price"), 0) for row in block_rows),
        "service_type": "",
        "is_subtotal": True,
    }


def estimate_total_row(option_key: str, total: float) -> Dict[str, Any]:
    return {
        "option": option_key,
        "group": "",
        "item_type": "",
        "selected_model": "",
        "component_type": "Estimate Total",
        "source_sheet": "",
        "line_number": "",
        "part_number": "",
        "smart_account_mandatory": "",
        "description": "",
        "group_name": "",
        "service_duration_months": "",
        "estimated_lead_time_days": "",
        "included_item": "",
        "quantity_per_unit": "",
        "quote_quantity": "",
        "total_quantity": "",
        "pricing_term": "",
        "list_price": "",
        "extended_list_price": "",
        "discount_percent": "Estimate Total",
        "selling_price": "",
        "extended_selling_price": total,
        "service_type": "",
        "is_estimate_total": True,
    }


def renumber_line_number(original: Any, block_number: int) -> str:
    text = str(original or "").strip()

    if not text:
        return f"{block_number}.0"

    if is_integer_line_number(text):
        return f"{block_number}.0"

    if "." in text:
        suffix = text.split(".", 1)[1]
        return f"{block_number}.{suffix}"

    return f"{block_number}.0"


def renumber_part_rows(rows: List[Dict[str, Any]], block_number: int) -> List[Dict[str, Any]]:
    result = []
    used = set()
    last_simple_decimal = -1

    for row in rows:
        item = dict(row)
        new_line_number = renumber_line_number(item.get("line_number"), block_number)
        parts = new_line_number.split(".")

        if len(parts) == 2 and parts[1].isdigit():
            decimal_value = int(parts[1])

            if new_line_number in used:
                decimal_value = last_simple_decimal + 1
                new_line_number = f"{block_number}.{decimal_value}"

            last_simple_decimal = max(last_simple_decimal, decimal_value)

        used.add(new_line_number)
        item["line_number"] = new_line_number
        result.append(item)

    return result


def build_bom(quote_data: Dict[str, Any]) -> Dict[str, Any]:
    quote = quote_data.get("quote", quote_data)
    quote_lines = quote.get("quote_lines", [])
    workbooks = load_detail_bom_workbooks()
    result: Dict[str, Any] = {"options": {}, "summary": {}}

    for option_key, option_label in OPTIONS:
        rows = []
        current_group = None
        block_number = 1

        for line in quote_lines:
            selected = (line.get("selected") or {}).get(option_key) or {}
            model = selected.get("model", "")
            quote_quantity = to_float(line.get("quantity"), 0)

            if not model or quote_quantity <= 0:
                continue

            group = str(line.get("group") or "Khác")

            if group != current_group:
                rows.append(group_header_row(option_key, group))
                current_group = group

            parts = find_bundle_parts(workbooks, model, quote_quantity, option_key, line)

            if not parts:
                parts = [fallback_selected_device_row(selected, quote_quantity, option_key, line)]

            parts = renumber_part_rows(parts, block_number)
            block_number += 1
            rows.extend(parts)
            rows.append(subtotal_row(option_key, group, parts))

        total = sum(
            to_float(row.get("extended_selling_price"), 0)
            for row in rows
            if not row.get("is_group_header") and not row.get("is_subtotal") and not row.get("is_estimate_total")
        )
        line_count = sum(
            1
            for row in rows
            if not row.get("is_group_header") and not row.get("is_subtotal") and not row.get("is_estimate_total")
        )
        rows.append(estimate_total_row(option_key, total))

        result["options"][option_key] = {
            "label": option_label,
            "rows": rows,
            "total": total,
            "line_count": line_count,
        }
        result["summary"][option_key] = {
            "label": option_label,
            "total": total,
            "line_count": line_count,
        }

    return result


def autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        width = 12

        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row, column=col).value

            if value is not None:
                width = max(width, min(len(str(value)) + 2, 48))

        ws.column_dimensions[get_column_letter(col)].width = width


def format_group_title(group: Any) -> str:
    return str(group or "").replace(" - ", "-").strip()


def write_rows_sheet(wb, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="969696")
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    group_border = Border(top=thin_side, bottom=thin_side)
    total_border = Border(right=thin_side, bottom=thin_side)
    header_font = Font(name="Helvetica", size=9, bold=True, color="000000")
    normal_font = Font(name="Helvetica", size=9, color="000000")
    bold_font = Font(name="Helvetica", size=9, bold=True, color="000000")
    total_font = Font(name="Helvetica", size=9, bold=True, color="0000FF")
    block_start_row = None
    subtotal_rows = []
    numeric_columns = {
        "total_quantity",
        "list_price",
        "extended_list_price",
        "discount_percent",
        "extended_selling_price",
    }

    for col, (_column_key, column_label) in enumerate(BOM_COLUMN_DEFINITIONS, start=1):
        cell = ws.cell(row=1, column=col, value=column_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 58

    for row_index, row in enumerate(rows, start=2):
        if row.get("is_group_header"):
            ws.cell(row=row_index, column=1, value=format_group_title(row.get("group", "")))
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(BOM_COLUMN_DEFINITIONS))
            ws.row_dimensions[row_index].height = 20

            cell = ws.cell(row=row_index, column=1)
            cell.font = bold_font
            cell.border = group_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in range(2, len(BOM_COLUMN_DEFINITIONS) + 1):
                merged_cell = ws.cell(row=row_index, column=col)
                merged_cell.border = group_border
            continue

        if row.get("is_subtotal") or row.get("is_estimate_total"):
            for col, (column_key, _column_label) in enumerate(BOM_COLUMN_DEFINITIONS, start=1):
                cell = ws.cell(row=row_index, column=col, value=row.get(column_key, ""))
                cell.font = total_font
                cell.alignment = Alignment(horizontal="right" if column_key in numeric_columns else "left", vertical="center")

            if row.get("is_subtotal"):
                if block_start_row and block_start_row <= row_index - 1:
                    ws.cell(row=row_index, column=BOM_COLUMN_KEYS.index("extended_list_price") + 1, value=f"=SUM(L{block_start_row}:L{row_index - 1})")
                    ws.cell(row=row_index, column=BOM_COLUMN_KEYS.index("extended_selling_price") + 1, value=f"=SUM(N{block_start_row}:N{row_index - 1})")

                subtotal_rows.append(row_index)
                block_start_row = None

            if row.get("is_estimate_total") and subtotal_rows:
                subtotal_formula = "+".join(f"N{subtotal_row}" for subtotal_row in subtotal_rows)
                ws.cell(row=row_index, column=BOM_COLUMN_KEYS.index("extended_selling_price") + 1, value=f"={subtotal_formula}")

            for col in range(1, len(BOM_COLUMN_DEFINITIONS) + 1):
                cell = ws.cell(row=row_index, column=col)
                cell.font = total_font
                cell.border = total_border
                cell.alignment = Alignment(horizontal="right", vertical="center")

            continue

        if block_start_row is None:
            block_start_row = row_index

        for col, (column_key, _column_label) in enumerate(BOM_COLUMN_DEFINITIONS, start=1):
            cell = ws.cell(row=row_index, column=col, value=row.get(column_key, ""))
            is_base_line = str(row.get("line_number", "")).strip().endswith(".0")
            cell.font = bold_font if column_key in {"line_number", "part_number"} and is_base_line else normal_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="right" if column_key in numeric_columns else "left",
                vertical="center",
                wrap_text=column_key in {"description", "service_type"},
            )

        quantity = to_float(row.get("total_quantity"), 0)
        list_price = to_float(row.get("list_price"), 0)
        extended_list_price = to_float(row.get("extended_list_price"), 0)

        if quantity > 0 and list_price > 0:
            term_factor = extended_list_price / (quantity * list_price)
            formula = f"=I{row_index}*K{row_index}"

            if abs(term_factor - 1) > 0.000001:
                formula = f"=I{row_index}*K{row_index}*{term_factor:.10g}"

            ws.cell(row=row_index, column=BOM_COLUMN_KEYS.index("extended_list_price") + 1, value=formula)
        else:
            ws.cell(row=row_index, column=BOM_COLUMN_KEYS.index("extended_list_price") + 1, value=extended_list_price)

        ws.cell(row=row_index, column=BOM_COLUMN_KEYS.index("extended_selling_price") + 1, value=f"=L{row_index}*(1-M{row_index}/100)")

    for col_name in ["list_price", "extended_list_price", "extended_selling_price"]:
        if col_name not in BOM_COLUMN_KEYS:
            continue

        col_index = BOM_COLUMN_KEYS.index(col_name) + 1

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_index).number_format = '#,##0.00'

    autosize_columns(ws)
    preferred_widths = {
        "A": 18,
        "B": 36,
        "C": 18,
        "D": 62,
        "E": 18,
        "F": 18,
        "G": 16,
        "H": 18,
        "I": 12,
        "J": 16,
        "K": 15,
        "L": 18,
        "M": 14,
        "N": 18,
        "O": 20,
    }
    for column_letter, width in preferred_widths.items():
        ws.column_dimensions[column_letter].width = width
    ws.freeze_panes = "A2"


def build_bom_excel(quote_data: Dict[str, Any], option_key: str | None = None) -> BytesIO:
    bom = build_bom(quote_data)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    selected_options = [
        (key, label)
        for key, label in OPTIONS
        if option_key is None or key == option_key
    ]

    if not selected_options:
        selected_options = OPTIONS

    for selected_key, option_label in selected_options:
        option = bom["options"][selected_key]
        write_rows_sheet(wb, option_label[:31], option["rows"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
