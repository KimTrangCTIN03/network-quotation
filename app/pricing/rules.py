from pathlib import Path
from typing import Any, Dict

import openpyxl

from app.catalog_engine import BOM_DIR, cell_value, is_integer_line_number, normalize_model, to_float


INDOOR_AP_POWER_INJECTOR = "AIR-PWRINJ7"
OUTDOOR_AP_POWER_INJECTOR = "IW-PWRINJ-60RGDMG"
OUTDOOR_AP_BRACKET = "AIR-MNT-VERT1"
OUTDOOR_AP_ACCESSORIES = [OUTDOOR_AP_POWER_INJECTOR, OUTDOOR_AP_BRACKET]
WIFI7_LICENSE_BUNDLE = "CISCO-NETWORK-SUB"


def read_bundle_subtotals(file_path: Path) -> Dict[str, float]:
    """Read one subtotal per base part from a BOM export sheet.

    The BOM files use the same layout as the Excel source: a numbered line
    starts a device/accessory block, then a later row marked "Subtotal" holds
    the bundle price. These subtotals are the raw building blocks for composed
    prices; the rules below decide when to add accessory/license blocks.
    """
    if not file_path.exists():
        return {}

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    result: Dict[str, float] = {}
    current_model = ""

    for row in range(2, ws.max_row + 1):
        line_number = cell_value(wb, ws, row, 1)
        model = normalize_model(cell_value(wb, ws, row, 2))

        if is_integer_line_number(line_number) and model:
            current_model = model
            continue

        subtotal_label = str(cell_value(wb, ws, row, 13) or "").strip().lower()

        if subtotal_label == "subtotal" and current_model:
            price = to_float(cell_value(wb, ws, row, 14), 0)

            if price > 0:
                result[current_model] = price

            current_model = ""

    return result


def read_named_bundle_subtotals(file_path: Path) -> Dict[str, float]:
    """Read subtotal blocks whose name is stored as a marker row.

    ISR1000 licenses are not keyed directly by a normal part row. In the source
    BOM they appear after marker text such as "License cho C1111-8P", followed
    by a subtotal row. This helper keeps that marker/subtotal relationship.
    """
    if not file_path.exists():
        return {}

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    result: Dict[str, float] = {}
    pending_name = ""

    for row in range(2, ws.max_row + 1):
        line_number = cell_value(wb, ws, row, 1)
        marker = str(cell_value(wb, ws, row, 3) or "").strip()

        if marker and not line_number:
            pending_name = marker

        subtotal_label = str(cell_value(wb, ws, row, 13) or "").strip().lower()

        if subtotal_label == "subtotal" and pending_name:
            price = to_float(cell_value(wb, ws, row, 14), 0)

            if price > 0:
                result[pending_name] = price

            pending_name = ""

    return result


def normalize_bundle_model(model: str) -> str:
    return normalize_model(model).replace("-ROW", "").replace("-RTG", "")


def add_isr1000_license_prices(prices: Dict[str, float]) -> None:
    file_path = BOM_DIR / "ISR1000.xlsx"
    base_prices = read_bundle_subtotals(file_path)
    named_subtotals = read_named_bundle_subtotals(file_path)

    for marker, license_price in named_subtotals.items():
        marker_lower = marker.lower()

        if not marker_lower.startswith("license cho "):
            continue

        # Mapping rule from Network Quotation Tool:
        # final ISR1000 price = base ISR hardware subtotal + the matching
        # "License cho <model>" subtotal block. The model name is extracted
        # from the marker itself, so adding a new ISR license block in BOM does
        # not require a hardcoded price in Python.
        model = normalize_model(marker[len("License cho "):])
        base_price = base_prices.get(model, 0)

        if model and base_price > 0 and license_price > 0:
            prices[model] = base_price + license_price


def add_ap_accessory_and_license_prices(prices: Dict[str, float]) -> None:
    file_path = BOM_DIR / "AP.xlsx"
    subtotals = read_bundle_subtotals(file_path)
    normalized_prices = {
        normalize_bundle_model(model): price
        for model, price in subtotals.items()
    }

    # AP mapping rule summary from the input mapping table:
    # - C9136 and CW916 indoor APs keep base price; injector is BOM-only.
    # - CW917 WiFi 7 APs use base price plus one Cisco Wireless license.
    # - C9124 outdoor APs and CW9163E use base price plus bracket; outdoor
    #   injector is BOM-only.
    # Accessory rows are kept in normalized_prices for calculations but skipped
    # from visible catalog output elsewhere.
    outdoor_bracket = normalized_prices.get(OUTDOOR_AP_BRACKET, 0)
    wifi7_license_unit = wifi7_license_price_per_ap(subtotals)

    for model, base_price in normalized_prices.items():
        if model in {INDOOR_AP_POWER_INJECTOR, WIFI7_LICENSE_BUNDLE, *OUTDOOR_AP_ACCESSORIES}:
            continue

        if model.startswith("CW917"):
            prices[model] = base_price + wifi7_license_unit
        elif model.startswith("C9124") or model == "CW9163E":
            prices[model] = base_price + outdoor_bracket
        elif model.startswith("C9136") or model.startswith("CW916"):
            prices[model] = base_price


def wifi7_license_price_per_ap(subtotals: Dict[str, float]) -> float:
    total = subtotals.get(WIFI7_LICENSE_BUNDLE, 0)

    if total <= 0:
        return 0

    # The BOM license bundle contains five repeated WiFi 7 AP license groups.
    return total / 5


def build_composed_list_prices() -> Dict[str, float]:
    prices: Dict[str, float] = {}
    add_isr1000_license_prices(prices)
    add_ap_accessory_and_license_prices(prices)
    return prices


def apply_composed_price_rules(prices: Dict[str, float]) -> Dict[str, float]:
    # Start with raw BOM subtotals, then override only the product lines that
    # need mapping-based final prices. This preserves ordinary devices as simple
    # subtotal prices while applying AP/ISR add-on logic consistently.
    result = dict(prices)
    result.update(build_composed_list_prices())
    return result


def rule_name_for_model(model: str) -> str:
    composed = build_composed_list_prices()

    if model in composed:
        return "bom_composed"

    return "bom_subtotal"
