from functools import lru_cache
import json
from pathlib import Path
from typing import Any, Dict, List
import re

import openpyxl
from openpyxl.utils import column_index_from_string


DATA_DIR = Path("data")

SPECS_FILE = DATA_DIR / "Devices Specs_240426.xlsx"
PRICE_FILE = DATA_DIR / "Cisco Unit List Price_240426.xlsx"
PRICE_OVERRIDE_FILE = DATA_DIR / "am_price_overrides.json"
BOM_DIR = DATA_DIR / "BOM"


SOURCE_BOM_SHEETS = [
    "C8000",
    "C9200",
    "C9300",
    "C1300",
    "EstimateDetails_CT167013261VV",
    "C9500",
    "ModularSwitch",
    "ACI",
    "ISR1000",
    "C8000_secure",
    "SFP",
    "AP",
]

SOURCE_BOM_FILES = [
    "C8000.xlsx",
    "C9200.xlsx",
    "C9300.xlsx",
    "C1300.xlsx",
    "C1200.xlsx",
    "C9500.xlsx",
    "ModularSwitch.xlsx",
    "ACI.xlsx",
    "ISR1000.xlsx",
    "C8000_secure.xlsx",
    "SFP.xlsx",
    "AP.xlsx",
]

C8000_SECURE_VARIANTS = {
    # C8000 Secure pricing is not a plain part-number lookup.
    # Each base chassis belongs to a size group, and only some commercial
    # variants are valid for that chassis. The tuple is:
    #   (<license size group>, [visible quote variants])
    # The visible model key becomes "<base model> (<variant>)".
    "C8355-G2": ("large", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8375-E-G2": ("large", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8550-G2": ("xlarge", ["Default Routing", "Adv Routing", "SDWAN"]),
    "C8570-G2": ("xlarge", ["Default Routing", "Adv Routing", "SDWAN"]),
    "C8455-G2": ("xlarge", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8475-G2": ("xlarge", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8235-E-G2": ("medium", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8231-E-G2": ("medium", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8235-G2": ("medium", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8231-G2": ("medium", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8130-G2": ("small", ["Default Routing", "Adv Routing"]),
    "C8131-G2": ("small", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8140-G2": ("small", ["Default Routing", "Adv Routing"]),
    "C8151-G2": ("small", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
    "C8161-G2": ("small", ["Default Routing", "Adv Routing", "SDWAN", "Full Adds On"]),
}

C8000_SECURE_LICENSE_PARTS = {
    # License add-on mapping by size group and variant.
    # Default Routing intentionally has no entry, so its final price is exactly
    # the base chassis subtotal. Other variants add the license part subtotals
    # found in C8000_secure.xlsx, never hardcoded numeric prices.
    "large": {
        "Adv Routing": ["LIC-ROS-L-A"],
        "SDWAN": ["LIC-CSWAN-L-A"],
        "Full Adds On": ["LIC-CSWAN-L-A", "LIC-SEC-L-T", "LIC-SEC-L-C", "LIC-SEC-L-M"],
    },
    "xlarge": {
        "Adv Routing": ["LIC-ROS-XL-A"],
        "SDWAN": ["LIC-CSWAN-XL-A"],
        "Full Adds On": ["LIC-CSWAN-XL-A", "LIC-SEC-XL-T", "LIC-SEC-XL-M", "LIC-SEC-XL-C"],
    },
    "medium": {
        "Adv Routing": ["LIC-ROS-M-A"],
        "SDWAN": ["LIC-CSWAN-M-A"],
        "Full Adds On": ["LIC-CSWAN-M-A", "LIC-SEC-M-T", "LIC-SEC-M-M", "LIC-SEC-M-C"],
    },
    "small": {
        "Adv Routing": ["LIC-ROS-S-A"],
        "SDWAN": ["LIC-CSWAN-S-A"],
        "Full Adds On": ["LIC-CSWAN-S-A", "LIC-SEC-S-T", "LIC-SEC-S-M", "LIC-SEC-S-C"],
    },
}


def normalize_model(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if not text:
        return ""

    text = text.rstrip("=")

    for suffix in ["-A", "-ROW", "-RTG"]:
        if text.endswith(suffix):
            text = text[: -len(suffix)]

    return text.strip()


def normalize_proposal_key(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip().replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    return text


def parse_google_export_formula(value: Any) -> Any:
    if value is None:
        return None

    if not isinstance(value, str):
        return value

    text = value.strip()

    if not text.startswith("="):
        return text

    m = re.search(r',\s*"([^"]*)"\s*\)\s*$', text)
    if m:
        return m.group(1)

    m = re.search(r',\s*(-?\d+(?:\.\d+)?)\s*\)\s*$', text)
    if m:
        number = float(m.group(1))
        return int(number) if number.is_integer() else number

    return None


def resolve_simple_ref(wb, formula: str) -> Any:
    if not isinstance(formula, str):
        return formula

    text = formula.strip()

    m = re.fullmatch(r"='?([^'!]+)'?!([A-Z]+)(\d+)", text)
    if not m:
        return None

    sheet_name = m.group(1)
    cell_ref = f"{m.group(2)}{m.group(3)}"

    if sheet_name not in wb.sheetnames:
        return None

    return parse_google_export_formula(wb[sheet_name][cell_ref].value)


def evaluate_simple_excel_formula(wb, ws, formula: str, seen: set[str] | None = None) -> Any:
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    expression = formula.strip()[1:].replace("$", "")
    seen = seen or set()

    try:
        while "ROUND(" in expression.upper():
            next_expression = replace_one_round_formula(wb, ws, expression, seen)
            if next_expression == expression:
                break
            expression = next_expression

        return evaluate_arithmetic_expression(wb, ws, expression, seen)
    except Exception:
        return None


def replace_one_round_formula(wb, ws, expression: str, seen: set[str]) -> str:
    start = expression.upper().rfind("ROUND(")
    if start < 0:
        return expression

    open_index = start + len("ROUND")
    depth = 0
    close_index = -1

    for index in range(open_index, len(expression)):
        char = expression[index]
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
            if depth == 0:
                close_index = index
                break

    if close_index < 0:
        return expression

    inner = expression[open_index + 1:close_index]
    comma_index = find_top_level_comma(inner)

    if comma_index < 0:
        return expression

    value_expression = inner[:comma_index]
    digits_expression = inner[comma_index + 1:]
    value = evaluate_arithmetic_expression(wb, ws, value_expression, seen)
    digits = evaluate_arithmetic_expression(wb, ws, digits_expression, seen)

    if value is None or digits is None:
        return expression

    rounded = round(value, int(digits))
    return f"{expression[:start]}{rounded}{expression[close_index + 1:]}"


def find_top_level_comma(expression: str) -> int:
    depth = 0

    for index, char in enumerate(expression):
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        elif char == "," and depth == 0:
            return index

    return -1


def evaluate_arithmetic_expression(wb, ws, expression: str, seen: set[str]) -> float | None:
    def cell_replacer(match):
        sheet_name = match.group("sheet")
        cell_ref = match.group("cell")
        if sheet_name:
            sheet_name = sheet_name.strip("'")
        target_ws = wb[sheet_name] if sheet_name else ws
        key = f"{target_ws.title}!{cell_ref}"

        if key in seen:
            raise ValueError("Circular formula reference")

        seen.add(key)
        value = cell_value(
            wb,
            target_ws,
            int(re.search(r"\d+", cell_ref).group(0)),
            column_index_from_string(re.search(r"[A-Z]+", cell_ref).group(0)),
        )
        seen.remove(key)

        return str(to_float(value, 0))

    expression = re.sub(
        r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_ ]+)!)?(?P<cell>[A-Z]{1,3}\d+)",
        cell_replacer,
        expression,
    )

    if not re.fullmatch(r"[0-9.\s+\-*/(),]+", expression):
        return None

    return float(eval(expression, {"__builtins__": {}}, {}))


def cell_value(wb, ws, row: int, col: int) -> Any:
    raw = ws.cell(row=row, column=col).value

    if isinstance(raw, str) and raw.startswith("="):
        ref_value = resolve_simple_ref(wb, raw)

        if ref_value is not None:
            return ref_value

        formula_value = evaluate_simple_excel_formula(wb, ws, raw)

        if formula_value is not None:
            return formula_value

    return parse_google_export_formula(raw)


def to_float(value: Any, default: float = 0) -> float:
    value = parse_google_export_formula(value)

    if value is None or value == "":
        return default

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")

    try:
        return float(text)
    except Exception:
        pass

    if re.fullmatch(r"[\d\.\+\-\*/\(\)\s]+", text):
        try:
            return float(eval(text, {"__builtins__": {}}))
        except Exception:
            return default

    return default


def is_integer_line_number(value: Any) -> bool:
    if value is None:
        return False

    text = str(value).strip()

    return re.fullmatch(r"\d+(\.0)?", text) is not None


def sheet_exists(file_path: Path, sheet_name: str) -> bool:
    if not file_path.exists():
        return False

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)

    return sheet_name in wb.sheetnames


def read_price_overrides() -> Dict[str, float]:
    if not PRICE_OVERRIDE_FILE.exists():
        return {}

    try:
        raw = json.loads(PRICE_OVERRIDE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

    result: Dict[str, float] = {}

    for model, price in raw.items():
        model_key = normalize_model(model)
        final_price = to_float(price, 0)

        if model_key and final_price > 0:
            result[model_key] = final_price

    return result


def read_3rd_sfp_price_map() -> Dict[str, float]:
    result: Dict[str, float] = {}

    if not sheet_exists(SPECS_FILE, "3rdSFP"):
        return result

    wb = openpyxl.load_workbook(SPECS_FILE, data_only=False)
    ws = wb["3rdSFP"]

    for row in range(2, ws.max_row + 1):
        model = normalize_model(cell_value(wb, ws, row, 1))
        list_price = to_float(cell_value(wb, ws, row, 2), 0)
        am_price = to_float(cell_value(wb, ws, row, 3), 0)
        final_price = to_float(cell_value(wb, ws, row, 4), 0)

        if final_price <= 0:
            final_price = am_price if am_price > 0 else list_price

        if model and final_price > 0:
            result[model] = final_price

    return result


def save_price_overrides(prices: Dict[str, float]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    normalized = {
        normalize_model(model): to_float(price, 0)
        for model, price in prices.items()
        if normalize_model(model) and to_float(price, 0) > 0
    }
    PRICE_OVERRIDE_FILE.write_text(
        json.dumps(normalized, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    clear_catalog_cache()


def clear_catalog_cache() -> None:
    read_price_map.cache_clear()
    load_catalogs.cache_clear()


@lru_cache(maxsize=1)
def read_price_map() -> Dict[str, float]:
    from app.pricing.catalog import read_final_price_map

    return read_final_price_map()


def read_cisco_tab_price_map() -> Dict[str, float]:
    if not PRICE_FILE.exists():
        return {}

    wb = openpyxl.load_workbook(PRICE_FILE, data_only=False)
    price_map: Dict[str, float] = {}

    if "Cisco" in wb.sheetnames:
        ws = wb["Cisco"]

        for row in range(2, ws.max_row + 1):
            model = normalize_model(cell_value(wb, ws, row, 1))

            if not model:
                continue

            list_price = to_float(cell_value(wb, ws, row, 2), 0)
            am_price = to_float(cell_value(wb, ws, row, 3), 0)
            final_price = to_float(cell_value(wb, ws, row, 4), 0)

            if final_price <= 0:
                final_price = am_price if am_price > 0 else list_price

            if final_price > 0:
                price_map[model] = final_price

    bom_price_map = read_price_map_from_bom_tabs(wb)

    for model, price in bom_price_map.items():
        if model not in price_map and price > 0:
            price_map[model] = price

    return price_map


def read_price_map_from_bom_files() -> Dict[str, float]:
    result: Dict[str, float] = {}

    if not BOM_DIR.exists():
        return result

    for file_name in SOURCE_BOM_FILES:
        file_path = BOM_DIR / file_name

        if not file_path.exists():
            continue

        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb.active

        if file_name == "C8000_secure.xlsx":
            read_c8000_secure_variant_prices(wb, ws, result)
        else:
            read_subtotal_bundle_prices(wb, ws, result)

    result.update(read_3rd_sfp_price_map())
    return result


def read_c8000_secure_variant_prices(wb, ws, result: Dict[str, float]) -> None:
    """Build C8000 Secure final prices from base chassis + license mapping.

    The Excel source has two kinds of rows that matter:
    1. Chassis blocks with a normal subtotal. These become base_prices.
    2. License part rows such as LIC-ROS-*, LIC-CSWAN-*, LIC-SEC-*.

    C8000_SECURE_VARIANTS decides which visible quote variants exist per base
    model. C8000_SECURE_LICENSE_PARTS decides which license part numbers are
    added for each variant. This mirrors the Excel formula/mapping behavior
    without importing final prices from "Cisco Unit List Price".
    """
    base_prices: Dict[str, float] = {}
    part_prices: Dict[str, float] = {}
    row = 2

    # First pass: collect every priced part row so license add-ons can be
    # referenced by part number later. Missing license part prices resolve to 0,
    # which keeps the function deterministic while making source data gaps easy
    # to spot through price comparison/debug output.
    for part_row in range(2, ws.max_row + 1):
        part_number = normalize_model(cell_value(wb, ws, part_row, 2))
        price = to_float(cell_value(wb, ws, part_row, 14), 0)

        if part_number and price > 0:
            part_prices[part_number] = price

    # Second pass: collect the base chassis subtotal for each C8000 Secure
    # model. The loop stops at the block subtotal or at the next numbered model
    # row, matching how the BOM groups rows visually.
    while row <= ws.max_row:
        model = normalize_model(cell_value(wb, ws, row, 2))

        if not is_integer_line_number(cell_value(wb, ws, row, 1)) or model not in C8000_SECURE_VARIANTS:
            row += 1
            continue

        subtotal = 0.0

        for scan_row in range(row + 1, ws.max_row + 1):
            next_line_number = cell_value(wb, ws, scan_row, 1)
            next_model = normalize_model(cell_value(wb, ws, scan_row, 2))
            subtotal_label = str(cell_value(wb, ws, scan_row, 13) or "").strip().lower()

            if subtotal_label == "subtotal":
                subtotal = to_float(cell_value(wb, ws, scan_row, 14), 0)
                row = scan_row
                break

            if is_integer_line_number(next_line_number) and next_model:
                break

        if subtotal > 0 and model not in base_prices:
            base_prices[model] = subtotal

        row += 1

    for model, base_price in base_prices.items():
        size, variants = C8000_SECURE_VARIANTS[model]
        license_parts = C8000_SECURE_LICENSE_PARTS[size]

        for variant in variants:
            # Final variant price:
            #   base chassis subtotal
            #   + sum(price of mapped license parts for this size/variant)
            # Default Routing has no mapped license parts, so sum(...) is 0.
            result[f"{model} ({variant})"] = base_price + sum(
                part_prices.get(normalize_model(part_number), 0)
                for part_number in license_parts.get(variant, [])
            )


def read_subtotal_bundle_prices(wb, ws, result: Dict[str, float]) -> None:
    base_models = []
    subtotals = []

    for row in range(2, ws.max_row + 1):
        line_number = cell_value(wb, ws, row, 1)
        item_name = cell_value(wb, ws, row, 2)

        if is_integer_line_number(line_number) and item_name:
            base_models.append(normalize_model(item_name))

        label_m = cell_value(wb, ws, row, 13)
        subtotal_n = cell_value(wb, ws, row, 14)

        if str(label_m or "").strip().lower() == "subtotal":
            price = to_float(subtotal_n, 0)

            if price > 0:
                subtotals.append(price)

    for model, price in zip(base_models, subtotals):
        if model and price > 0:
            result[model] = price


def read_price_map_from_bom_tabs(wb) -> Dict[str, float]:
    result: Dict[str, float] = {}

    for sheet_name in SOURCE_BOM_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        read_aggregated_sheet_prices(wb, ws, result)
        read_subtotal_bundle_prices(wb, ws, result)

    return result


def read_aggregated_sheet_prices(wb, ws, result: Dict[str, float]) -> None:
    header_map = {}

    for col in range(1, ws.max_column + 1):
        header = cell_value(wb, ws, 1, col)

        if header is None:
            continue

        header_map[str(header).strip().lower()] = col

    possible_device_cols = [
        header_map.get("devices"),
        header_map.get("device"),
        header_map.get("item name"),
    ]

    possible_price_cols = [
        header_map.get("final price"),
        header_map.get("price base"),
        header_map.get("list price"),
    ]

    device_cols = [c for c in possible_device_cols if c]
    price_cols = [c for c in possible_price_cols if c]

    if not device_cols or not price_cols:
        return

    device_col = device_cols[0]
    price_col = price_cols[-1]

    for row in range(2, ws.max_row + 1):
        model = normalize_model(cell_value(wb, ws, row, device_col))
        price = to_float(cell_value(wb, ws, row, price_col), 0)

        if model and price > 0:
            result[model] = price


CLASS_VALUES = {"Low End", "Mid Range", "High End"}


def normalize_device_class(value: Any) -> str:
    text = str(value or "").strip()

    for class_name in CLASS_VALUES:
        if text.lower() == class_name.lower():
            return class_name

    return ""


def find_class_in_specs(specs: Dict[str, Any]) -> str:
    for key, value in specs.items():
        if "class" in str(key).lower():
            class_name = normalize_device_class(value)

            if class_name:
                return class_name

    return ""


def infer_wifi_class(specs: Dict[str, Any]) -> str:
    technology = str(specs.get("Công nghệ WiFi (WiFi6/WiFi7)") or "").upper()

    if "WIFI7" in technology:
        return "High End"

    if "WIFI6" in technology:
        return "Mid Range"

    return ""


def read_specs_matrix_sheet(sheet_name: str) -> List[Dict[str, Any]]:
    if not SPECS_FILE.exists():
        return []

    wb = openpyxl.load_workbook(SPECS_FILE, data_only=False)

    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    price_map = read_price_map()

    row_labels: Dict[int, str] = {}

    for row in range(1, min(ws.max_row, 150) + 1):
        label = cell_value(wb, ws, row, 1)

        if label:
            row_labels[row] = str(label).strip()

    devices: List[Dict[str, Any]] = []

    for col in range(2, ws.max_column + 1):
        model = normalize_model(cell_value(wb, ws, 1, col))

        if not model or model == "Requirement" or model.startswith("<openpyxl."):
            continue

        specs = {}
        label_counts: Dict[str, int] = {}

        for row, label in row_labels.items():
            value = cell_value(wb, ws, row, col)
            label_counts[label] = label_counts.get(label, 0) + 1
            spec_label = label if label_counts[label] == 1 else f"{label} #{label_counts[label]}"
            specs[spec_label] = value

        device_class = find_class_in_specs(specs)

        if sheet_name == "WiFi":
            device_class = device_class or infer_wifi_class(specs)
            if device_class:
                specs["Access Point Class"] = device_class

        devices.append({
            "model": model,
            "sheet": sheet_name,
            "class": device_class,
            "price": price_map.get(model, 0),
            "specs": specs,
        })

    return devices


SWITCH_SPEC_KEYS = {
    "class": "Switch Class",
    "bandwidth": "Switching Bandwidth - Full Duplex (Gbps)",
    "forwarding": "Forwarding Capacity (Mpps)",
    "1g_rj45": "Sá»‘ lÆ°á»£ng cá»•ng 1GE Ä‘á»“ng",
    "1g_sfp": "Sá»‘ lÆ°á»£ng cá»•ng 1GE SFP",
    "10g_rj45": "Sá»‘ lÆ°á»£ng cá»•ng 10GE Ä‘á»“ng",
    "10g_sfp": "Sá»‘ lÆ°á»£ng cá»•ng 10GE quang",
    "100g": "Sá»‘ lÆ°á»£ng cá»•ng 100GE",
    "stacking": "Stacking (Y/N)",
    "poe": "PoE (Y/N)",
}


def infer_switch_class_from_model(model: str) -> str:
    text = model.upper()

    if text.startswith(("C1200", "C1300", "C9200")):
        return "Low End"
    if text.startswith(("C9300", "C9400", "C9500", "N9K-C93")):
        return "Mid Range"
    if text.startswith(("C9600", "C9500X", "N9K-C95", "N9K-C936", "N9K-C933")):
        return "High End"
    return ""


def looks_like_switch_model(model: str) -> bool:
    text = model.upper()
    return text.startswith(("C1200", "C1300", "C9200", "C9300", "C9400", "C9500", "C9600", "N9K-"))


def infer_switch_ports_from_model(model: str) -> Dict[str, float]:
    text = model.upper()
    ports = {
        "1g_rj45": 0.0,
        "1g_sfp": 0.0,
        "10g_rj45": 0.0,
        "10g_sfp": 0.0,
        "100g": 0.0,
    }

    if match := re.search(r"-(\d+)([A-Z]+)", text):
        count = to_float(match.group(1), 0)
        suffix = match.group(2)

        if "D" in suffix or "C" in suffix:
            ports["100g"] = max(ports["100g"], count)
        elif "Y" in suffix or "S" in suffix or "X" in suffix:
            ports["10g_sfp"] = max(ports["10g_sfp"], count)
        elif "T" in suffix:
            ports["1g_rj45"] = max(ports["1g_rj45"], count)
        elif "P" in suffix or "U" in suffix:
            ports["1g_rj45"] = max(ports["1g_rj45"], count)

    for count_text, speed_text in re.findall(r"-(\d+)([GX])", text):
        count = to_float(count_text, 0)
        if speed_text == "X":
            ports["10g_sfp"] = max(ports["10g_sfp"], count)
        elif speed_text == "G":
            ports["1g_sfp"] = max(ports["1g_sfp"], count)

    if text.startswith("N9K-C93108TC") or text.startswith("N9K-C93216TC"):
        ports["10g_rj45"] = max(ports["10g_rj45"], 48)
        ports["100g"] = max(ports["100g"], 6)
    elif text.startswith("N9K-C93180YC") or text.startswith("N9K-C93240YC") or text.startswith("N9K-C93360YC"):
        ports["10g_sfp"] = max(ports["10g_sfp"], 48)
        ports["100g"] = max(ports["100g"], 6)
    elif text.startswith("N9K-C936") or text.startswith("N9K-X97"):
        ports["100g"] = max(ports["100g"], 32)

    return ports


def synthetic_switch_device(model: str, price: float) -> Dict[str, Any]:
    model_key = normalize_model(model)
    ports = infer_switch_ports_from_model(model_key)
    specs = {
        SWITCH_SPEC_KEYS["class"]: infer_switch_class_from_model(model_key),
        SWITCH_SPEC_KEYS["bandwidth"]: 0,
        SWITCH_SPEC_KEYS["forwarding"]: 0,
        SWITCH_SPEC_KEYS["1g_rj45"]: ports["1g_rj45"],
        SWITCH_SPEC_KEYS["1g_sfp"]: ports["1g_sfp"],
        SWITCH_SPEC_KEYS["10g_rj45"]: ports["10g_rj45"],
        SWITCH_SPEC_KEYS["10g_sfp"]: ports["10g_sfp"],
        SWITCH_SPEC_KEYS["100g"]: ports["100g"],
        SWITCH_SPEC_KEYS["stacking"]: "Y" if model_key.upper().startswith(("C9200", "C9300", "C9500", "C9600")) else "",
        SWITCH_SPEC_KEYS["poe"]: "Y" if re.search(r"-\d+[A-Z]*P", model_key.upper()) else "",
        "selection_source": "bom_price_catalog",
    }
    sheet = "NexusSwitch" if model_key.upper().startswith("N9K-") else "SwitchCampus"

    return {
        "model": model_key,
        "sheet": sheet,
        "class": infer_switch_class_from_model(model_key),
        "price": price,
        "specs": specs,
    }


def merge_synthetic_switches(devices: List[Dict[str, Any]], sheet_name: str, prices: Dict[str, float]) -> List[Dict[str, Any]]:
    by_model = {normalize_model(device.get("model")): dict(device) for device in devices if normalize_model(device.get("model"))}

    for model, price in prices.items():
        model_key = normalize_model(model)
        if not model_key or model_key in by_model or not looks_like_switch_model(model_key):
            continue

        synthetic = synthetic_switch_device(model_key, price)
        if synthetic["sheet"] == sheet_name:
            by_model[model_key] = synthetic

    return sorted(
        by_model.values(),
        key=lambda device: (
            to_float(device.get("price"), 0) <= 0,
            to_float(device.get("price"), 0),
            str(device.get("model") or ""),
        ),
    )


def infer_sfp_speed(model: str, description: str = "") -> float:
    text = f"{model} {description}".upper()

    if "100G" in text:
        return 100

    if "40G" in text:
        return 40

    if "25G" in text:
        return 25

    if "10G" in text:
        return 10

    if "1G" in text or "GLC" in text:
        return 1

    return 0


def infer_sfp_distance(model: str, description: str = "") -> float:
    text = f"{model} {description}".upper()

    explicit_km = re.search(r"(\d+(?:\.\d+)?)\s*KM", text)
    if explicit_km:
        return to_float(explicit_km.group(1), 0)

    if "80KM" in text or "ZR" in text:
        return 80

    if "40KM" in text or "ER" in text:
        return 40

    if "10KM" in text or "LR" in text or "LH" in text:
        return 10

    if "SR" in text:
        return 0.3

    return 0


def looks_like_sfp_model(model: str) -> bool:
    text = str(model or "").upper().strip()

    if not text:
        return False

    keywords = [
        "SFP",
        "QSFP",
        "GLC",
        "FET",
        "CWDM",
        "DWDM",
        "X2-",
        "XFP",
        "CFP",
        "CVR-QSFP",
        "QDD",
        "FICER",
    ]

    return any(k in text for k in keywords)


def infer_sfp_class(model: str) -> str:
    text = model.upper()

    if "FICER" in text:
        return "Low End"

    if "QSFP" in text or "SFP" in text or "GLC" in text:
        return "High End"

    return "Mid Range"


def read_sfp_catalog() -> List[Dict[str, Any]]:
    price_map = read_price_map()
    devices_by_model: Dict[str, Dict[str, Any]] = {}

    def add_sfp_device(model: Any, description: str = "", price: float = 0):
        model_name = normalize_model(model)

        if not model_name:
            return

        if not looks_like_sfp_model(model_name):
            return

        final_price = price_map.get(model_name, price)

        speed = infer_sfp_speed(model_name, description)
        distance = infer_sfp_distance(model_name, description)

        devices_by_model[model_name] = {
            "model": model_name,
            "sheet": "SFP",
            "class": infer_sfp_class(model_name),
            "price": final_price,
            "speed": speed,
            "distance": distance,
            "specs": {
                "description": description,
                "speed": speed,
                "distance": distance,
            }
        }

    if sheet_exists(SPECS_FILE, "SFP"):
        wb = openpyxl.load_workbook(SPECS_FILE, data_only=False)
        ws = wb["SFP"]

        for col in range(2, ws.max_column + 1):
            model = normalize_model(cell_value(wb, ws, 1, col))

            if looks_like_sfp_model(model):
                speed = to_float(cell_value(wb, ws, 2, col), 0)
                distance = to_float(cell_value(wb, ws, 3, col), 0)
                sfp_class = str(cell_value(wb, ws, 4, col) or "").strip()
                price = price_map.get(model, to_float(cell_value(wb, ws, 5, col), 0))

                devices_by_model[model] = {
                    "model": model,
                    "sheet": "SFP",
                    "class": sfp_class or infer_sfp_class(model),
                    "price": price,
                    "speed": speed or infer_sfp_speed(model),
                    "distance": distance or infer_sfp_distance(model),
                    "specs": {
                        "speed": speed or infer_sfp_speed(model),
                        "distance": distance or infer_sfp_distance(model),
                    }
                }

        for row in range(2, ws.max_row + 1):
            item_name = cell_value(wb, ws, row, 2)

            if not item_name:
                continue

            model = normalize_model(item_name)

            if not looks_like_sfp_model(model):
                continue

            description_parts = []

            for col in range(3, min(ws.max_column, 8) + 1):
                v = cell_value(wb, ws, row, col)

                if v:
                    description_parts.append(str(v))

            description = " ".join(description_parts)
            price = price_map.get(model, 0)

            add_sfp_device(model, description, price)

    for model, price in price_map.items():
        if looks_like_sfp_model(model):
            add_sfp_device(model, "", price)

    return sorted(
        list(devices_by_model.values()),
        key=lambda x: (
            x.get("speed", 0) or 0,
            x.get("price", 0) or 0,
            x.get("model", "")
        )
    )


@lru_cache(maxsize=1)
def load_catalogs() -> Dict[str, Any]:
    prices = read_price_map()
    return {
        "prices": prices,
        "routers": read_specs_matrix_sheet("Router"),
        "switches": merge_synthetic_switches(read_specs_matrix_sheet("SwitchCampus"), "SwitchCampus", prices),
        "modular_switches": read_specs_matrix_sheet("ModularSwitch"),
        "nexus_switches": merge_synthetic_switches(read_specs_matrix_sheet("NexusSwitch"), "NexusSwitch", prices),
        "wifi": read_specs_matrix_sheet("WiFi"),
        "sfps": read_sfp_catalog(),
    }


def debug_catalog_summary() -> Dict[str, Any]:
    catalogs = load_catalogs()

    return {
        "price_count": len(catalogs["prices"]),
        "router_count": len(catalogs["routers"]),
        "switch_count": len(catalogs["switches"]),
        "modular_switch_count": len(catalogs["modular_switches"]),
        "nexus_switch_count": len(catalogs["nexus_switches"]),
        "wifi_count": len(catalogs["wifi"]),
        "sfp_count": len(catalogs["sfps"]),
        "sample_prices": list(catalogs["prices"].items())[:10],
    }


def compare_price_map_with_cisco_tab(limit: int = 50) -> Dict[str, Any]:
    from app.pricing.catalog import read_list_price_map

    current = read_list_price_map()
    cisco = read_cisco_tab_list_price_map()
    mismatches = []
    missing_in_current = []
    missing_in_cisco = []

    for model, cisco_price in cisco.items():
        current_price = current.get(model)

        if current_price is None:
            missing_in_current.append({"model": model, "cisco_price": cisco_price})
            continue

        if round(float(current_price), 2) != round(float(cisco_price), 2):
            mismatches.append({
                "model": model,
                "current_price": current_price,
                "cisco_price": cisco_price,
                "delta": current_price - cisco_price,
            })

    for model, current_price in current.items():
        if model not in cisco:
            missing_in_cisco.append({"model": model, "current_price": current_price})

    return {
        "current_count": len(current),
        "cisco_count": len(cisco),
        "matched_count": len(cisco) - len(missing_in_current) - len(mismatches),
        "mismatch_count": len(mismatches),
        "missing_in_current_count": len(missing_in_current),
        "missing_in_cisco_count": len(missing_in_cisco),
        "mismatches": mismatches[:limit],
        "missing_in_current": missing_in_current[:limit],
        "missing_in_cisco": missing_in_cisco[:limit],
    }


def read_cisco_tab_list_price_map() -> Dict[str, float]:
    if not PRICE_FILE.exists():
        return {}

    wb = openpyxl.load_workbook(PRICE_FILE, data_only=False)
    result: Dict[str, float] = {}

    if "Cisco" not in wb.sheetnames:
        return result

    ws = wb["Cisco"]

    for row in range(2, ws.max_row + 1):
        model = normalize_model(cell_value(wb, ws, row, 1))
        price = to_float(cell_value(wb, ws, row, 2), 0)

        if model and price > 0:
            result[model] = price

    return result
