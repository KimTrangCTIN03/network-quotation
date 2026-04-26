from io import BytesIO
from typing import Any, Dict, List
import unicodedata

import openpyxl

from app.catalog_engine import (
    clear_catalog_cache,
    normalize_model,
    read_price_map,
    to_float,
)
from app.pricing.catalog import build_price_catalog, normalize_vendor
from app.pricing.storage import save_am_price_entry, save_custom_price_entry


MODEL_HEADERS = [
    "model",
    "part number",
    "part_number",
    "part no",
    "part no.",
    "sku",
    "device",
    "devices",
    "item name",
    "item",
    "ma thiet bi",
    "m? thi?t b?",
    "ten thiet bi",
    "part",
    "partnumber",
]
DESCRIPTION_HEADERS = ["description", "product description", "desc"]
QTY_HEADERS = [
    "qty",
    "quantity",
    "quote qty",
    "quote quantity",
    "total_quantity",
    "total qty",
    "total quantity",
]
AM_PRICE_HEADERS = [
    "final price",
    "am price",
    "selling price",
    "unit_price",
    "unit price",
    "gia am",
    "gi? am",
    "gia ban",
]
LIST_PRICE_HEADERS = [
    "list_price",
    "list price",
    "unit list price",
    "list",
    "gia list",
    "gia niem yet",
    "list gia",
]
PRICE_HEADERS = AM_PRICE_HEADERS + LIST_PRICE_HEADERS
IMPORT_REQUIRED_LABELS = {
    "model": "Device/Mã thiết bị",
    "list_price": "List Price",
    "am_price": "AM Price/Giá AM",
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("\n", " ")
    text = text.replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = " ".join(text.replace("_", " ").split())
    return text


def find_header_row(ws) -> tuple[int, Dict[str, int]]:
    best_row = 1
    best_headers: Dict[str, int] = {}
    best_score = 0

    for row in range(1, min(ws.max_row, 30) + 1):
        headers = {}

        for col in range(1, ws.max_column + 1):
            header = normalize_header(ws.cell(row=row, column=col).value)

            if header:
                headers[header] = col

        score = 0
        if any(header in MODEL_HEADERS for header in headers):
            score += 2
        if any(header in QTY_HEADERS for header in headers):
            score += 2
        if any(header in PRICE_HEADERS for header in headers):
            score += 1

        if score > best_score:
            best_row = row
            best_headers = headers
            best_score = score

    return best_row, best_headers


def first_col(headers: Dict[str, int], names: List[str]) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]

    return None


def parse_bom_rows(content: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    rows: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws)
        model_col = first_col(headers, MODEL_HEADERS)

        if not model_col:
            continue

        description_col = first_col(headers, DESCRIPTION_HEADERS)
        quantity_col = first_col(headers, QTY_HEADERS)
        am_price_col = first_col(headers, AM_PRICE_HEADERS)
        list_price_col = first_col(headers, LIST_PRICE_HEADERS)
        generic_price_col = first_col(headers, PRICE_HEADERS)

        for row_index in range(header_row + 1, ws.max_row + 1):
            raw_model = ws.cell(row=row_index, column=model_col).value
            model = normalize_model(raw_model)

            if not model:
                continue

            quantity = to_float(
                ws.cell(row=row_index, column=quantity_col).value if quantity_col else 1,
                1,
            )

            if quantity <= 0:
                quantity = 1

            rows.append({
                "sheet": ws.title,
                "row": row_index,
                "model": model,
                "description": (
                    str(ws.cell(row=row_index, column=description_col).value or "").strip()
                    if description_col
                    else ""
                ),
                "quantity": quantity,
                "input_price": (
                    to_float(ws.cell(row=row_index, column=am_price_col).value, 0)
                    if am_price_col
                    else to_float(ws.cell(row=row_index, column=generic_price_col).value, 0)
                    if generic_price_col and generic_price_col != list_price_col
                    else 0
                ),
                "list_price": (
                    to_float(ws.cell(row=row_index, column=list_price_col).value, 0)
                    if list_price_col
                    else 0
                ),
            })

    return rows


def parse_price_import_rows(content: bytes) -> tuple[List[Dict[str, Any]], List[str]]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    rows: List[Dict[str, Any]] = []
    missing_templates: List[str] = []

    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws)
        model_col = first_col(headers, MODEL_HEADERS)
        list_price_col = first_col(headers, LIST_PRICE_HEADERS)
        am_price_col = first_col(headers, AM_PRICE_HEADERS)

        if not model_col or not list_price_col or not am_price_col:
            missing = []
            if not model_col:
                missing.append(IMPORT_REQUIRED_LABELS["model"])
            if not list_price_col:
                missing.append(IMPORT_REQUIRED_LABELS["list_price"])
            if not am_price_col:
                missing.append(IMPORT_REQUIRED_LABELS["am_price"])

            missing_templates.append(f"{ws.title}: thiếu {', '.join(missing)}")
            continue

        for row_index in range(header_row + 1, ws.max_row + 1):
            model = normalize_model(ws.cell(row=row_index, column=model_col).value)

            if not model:
                continue

            rows.append({
                "sheet": ws.title,
                "row": row_index,
                "model": model,
                "description": "",
                "quantity": 1,
                "input_price": to_float(ws.cell(row=row_index, column=am_price_col).value, 0),
                "list_price": to_float(ws.cell(row=row_index, column=list_price_col).value, 0),
            })

    return rows, missing_templates


def quote_bom(content: bytes, filename: str = "") -> Dict[str, Any]:
    price_map = read_price_map()
    rows = parse_bom_rows(content)
    quoted_rows = []
    total = 0.0
    matched = 0

    for row in rows:
        model = row["model"]
        system_price = price_map.get(model, 0)
        unit_price = system_price if system_price > 0 else row.get("input_price", 0)
        amount = unit_price * row["quantity"]
        total += amount

        if system_price > 0:
            matched += 1

        quoted_rows.append({
            **row,
            "unit_price": unit_price,
            "amount": amount,
            "price_source": "system" if system_price > 0 else "bom" if unit_price > 0 else "missing",
        })

    return {
        "filename": filename,
        "summary": {
            "line_count": len(quoted_rows),
            "matched_price_count": matched,
            "missing_price_count": len(quoted_rows) - matched,
            "total": total,
        },
        "rows": quoted_rows,
    }


def build_existing_price_map() -> Dict[str, Dict[str, float]]:
    return {
        f"{row['vendor'].lower()}::{row['model']}": {
            "list_price": to_float(row.get("list_price"), 0),
            "am_price": to_float(row.get("am_price"), 0),
        }
        for row in build_price_catalog()
    }


def find_import_conflicts(rows: List[Dict[str, Any]], vendor_key: str) -> List[Dict[str, Any]]:
    existing = build_existing_price_map()
    seen: Dict[str, Dict[str, Any]] = {}
    conflicts = []

    for row in rows:
        model = row["model"]
        am_price = to_float(row.get("input_price"), 0)
        list_price = to_float(row.get("list_price"), 0)

        if not model or (am_price <= 0 and list_price <= 0):
            continue

        key = f"{vendor_key.lower()}::{model}"
        old = existing.get(key)

        list_price_conflict = list_price > 0 and old and old["list_price"] > 0 and old["list_price"] != list_price
        am_price_conflict = am_price > 0 and old and (
            (old["am_price"] > 0 and old["am_price"] != am_price)
            or (old["am_price"] <= 0 and old["list_price"] > 0 and old["list_price"] != am_price)
        )

        if old and (list_price_conflict or am_price_conflict):
            conflicts.append({
                "type": "existing",
                "vendor": vendor_key,
                "model": model,
                "old_list_price": old["list_price"],
                "new_list_price": list_price,
                "old_am_price": old["am_price"],
                "new_am_price": am_price,
                "sheet": row["sheet"],
                "row": row["row"],
            })

        if key in seen:
            previous = seen[key]
            conflicts.append({
                "type": "file_duplicate",
                "vendor": vendor_key,
                "model": model,
                "old_list_price": to_float(previous.get("list_price"), 0),
                "new_list_price": list_price,
                "old_am_price": to_float(previous.get("input_price"), 0),
                "new_am_price": am_price,
                "sheet": row["sheet"],
                "row": row["row"],
            })

        seen[key] = row

    return conflicts


def import_prices_from_bom(
    content: bytes,
    filename: str = "",
    vendor: str = "",
    confirm_overwrite: bool = False,
) -> Dict[str, Any]:
    rows, template_errors = parse_price_import_rows(content)
    vendor_key = normalize_vendor(vendor)
    imported = []
    skipped = 0
    am_price_count = 0
    list_price_count = 0

    if template_errors and not rows:
        return {
            "filename": filename,
            "imported_count": 0,
            "skipped_count": 0,
            "template_error": True,
            "message": "File import phải có đủ 3 cột: Device/Mã thiết bị, List Price, AM Price/Giá AM.",
            "details": template_errors,
        }

    conflicts = find_import_conflicts(rows, vendor_key)

    if conflicts and not confirm_overwrite:
        return {
            "filename": filename,
            "requires_confirmation": True,
            "conflict_count": len(conflicts),
            "conflicts": conflicts[:50],
            "parsed_count": len(rows),
        }

    for row in rows:
        model = row["model"]
        am_price = to_float(row.get("input_price"), 0)
        list_price = to_float(row.get("list_price"), 0)

        if not model or (am_price <= 0 and list_price <= 0):
            skipped += 1
            continue

        if list_price > 0:
            save_custom_price_entry(vendor_key, model, list_price)
            list_price_count += 1

        if am_price > 0:
            save_am_price_entry(vendor_key, model, am_price)
            am_price_count += 1

        imported.append({
            "vendor": vendor_key,
            "model": model,
            "am_price": am_price,
            "list_price": list_price,
            "sheet": row["sheet"],
            "row": row["row"],
        })

    clear_catalog_cache()

    return {
        "filename": filename,
        "imported_count": len(imported),
        "am_price_count": am_price_count,
        "list_price_count": list_price_count,
        "skipped_count": skipped,
        "template_warnings": template_errors,
        "rows": imported,
    }
